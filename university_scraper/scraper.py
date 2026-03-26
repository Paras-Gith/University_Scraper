import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.hyperlink import Hyperlink
import time
import os

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

SHIKSHA_URLS = {
    "U001": "https://www.shiksha.com/university/iit-bombay-indian-institute-of-technology-mumbai-54212/courses",
    "U002": "https://www.shiksha.com/university/iisc-bangalore-indian-institute-of-science-9079/courses",
    "U003": "https://www.shiksha.com/university/university-of-delhi-9181/courses",
    "U004": "https://www.shiksha.com/university/jawaharlal-nehru-university-delhi-4225/courses",
    "U005": "https://www.shiksha.com/university/iit-delhi-indian-institute-of-technology-53938/courses",
}

def scrape_shiksha(university_id):
    url = SHIKSHA_URLS.get(university_id)
    if not url:
        return []

    try:
        print(f"  [SCRAPING] Shiksha for {university_id} ...")
        resp = requests.get(url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")

        courses = []
        selectors = [
            "div.college-course-card",
            "div.course-list-item",
            "div[class*='course-card']",
            "li[class*='course']",
            "div.courseCard",
            "tr.course-row",
        ]

        cards = []
        for sel in selectors:
            cards = soup.select(sel)
            if cards:
                print(f"    Found {len(cards)} cards using: {sel}")
                break

        for card in cards[:7]:
            name_el  = card.select_one("h3, h4, [class*='name'], [class*='title']")
            level_el = card.select_one("[class*='level'], [class*='degree']")
            dur_el   = card.select_one("[class*='duration'], [class*='year']")
            fee_el   = card.select_one("[class*='fee'], [class*='price']")

            name = name_el.get_text(strip=True) if name_el else None
            if name and len(name) > 4:
                courses.append({
                    "course_name": name,
                    "level":       level_el.get_text(strip=True) if level_el else "Not Available",
                    "discipline":  "Not Available",
                    "duration":    dur_el.get_text(strip=True)   if dur_el   else "Not Available",
                    "fees":        fee_el.get_text(strip=True)   if fee_el   else "Not Available",
                    "eligibility": "Not Available",
                })

        if courses:
            print(f"    SUCCESS: {len(courses)} courses scraped")
        else:
            print(f"    No courses parsed — will use fallback")

        return courses

    except Exception as e:
        print(f"    FAILED ({e}) — will use fallback")
        return []


FALLBACK_DATA = {
    "U001": {
        "name":    "Indian Institute of Technology Bombay (IIT Bombay)",
        "country": "India",
        "city":    "Mumbai",
        "website": "https://www.iitb.ac.in",
        "courses": [
            ("B.Tech in Computer Science & Engineering",   "Bachelor's", "Computer Science",       "4 Years",   "₹2,28,650/year",    "JEE Advanced, Class 12 with 75%+"),
            ("B.Tech in Electrical Engineering",           "Bachelor's", "Electrical Engineering", "4 Years",   "₹2,28,650/year",    "JEE Advanced, Class 12 with 75%+"),
            ("B.Tech in Mechanical Engineering",           "Bachelor's", "Mechanical Engineering", "4 Years",   "₹2,28,650/year",    "JEE Advanced, Class 12 with 75%+"),
            ("M.Tech in Artificial Intelligence",          "Master's",   "AI / Computer Science",  "2 Years",   "₹59,100/semester",  "GATE qualified, B.Tech in CS/EE"),
            ("M.Tech in Data Science",                     "Master's",   "Data Science",           "2 Years",   "₹59,100/semester",  "GATE qualified, B.Tech in CS/Math"),
            ("M.Sc in Applied Statistics & Informatics",  "Master's",   "Mathematics",            "2 Years",   "₹14,800/semester",  "JAM qualified, B.Sc in Math/Stats"),
            ("PhD in Computer Science & Engineering",      "PhD",        "Computer Science",       "4-6 Years", "₹14,800/semester",  "M.Tech/M.Sc + GATE/NET"),
        ]
    },
    "U002": {
        "name":    "Indian Institute of Science (IISc)",
        "country": "India",
        "city":    "Bangalore",
        "website": "https://www.iisc.ac.in",
        "courses": [
            ("B.S. in Research (Physics)",                 "Bachelor's", "Physics",                "4 Years",   "₹37,500/semester",  "JEE Advanced / KVPY / CUET"),
            ("B.S. in Research (Chemistry)",               "Bachelor's", "Chemistry",              "4 Years",   "₹37,500/semester",  "JEE Advanced / KVPY / CUET"),
            ("M.Tech in Computational Science",            "Master's",   "Computer Science/Math",  "2 Years",   "₹50,000/semester",  "GATE qualified"),
            ("M.Tech in AI & Machine Learning",            "Master's",   "Artificial Intelligence","2 Years",   "₹50,000/semester",  "GATE in CS/EC/EE"),
            ("M.Sc in Biochemistry",                       "Master's",   "Life Sciences",          "2 Years",   "₹25,000/semester",  "JAM qualified, B.Sc in Science"),
            ("M.Sc in Physics",                            "Master's",   "Physics",                "2 Years",   "₹25,000/semester",  "JAM qualified, B.Sc in Physics"),
            ("PhD in Electrical Communication Engg.",      "PhD",        "Electronics & Comm.",    "5-6 Years", "₹14,400/semester",  "GATE / UGC-NET / CSIR-NET"),
        ]
    },
    "U003": {
        "name":    "University of Delhi (DU)",
        "country": "India",
        "city":    "New Delhi",
        "website": "https://www.du.ac.in",
        "courses": [
            ("B.A. (Hons) Economics",                      "Bachelor's", "Economics",              "3 Years",   "₹13,980/year",      "Class 12 with 60%+, CUET-UG"),
            ("B.Com (Hons)",                               "Bachelor's", "Commerce",               "3 Years",   "₹12,615/year",      "Class 12 with Commerce, CUET-UG"),
            ("B.Sc (Hons) Computer Science",               "Bachelor's", "Computer Science",       "3 Years",   "₹14,580/year",      "Class 12 with Math, CUET-UG"),
            ("M.A. in Political Science",                  "Master's",   "Political Science",      "2 Years",   "₹14,580/year",      "Bachelor's Degree, CUET-PG"),
            ("M.Sc in Computer Science",                   "Master's",   "Computer Science",       "2 Years",   "₹16,380/year",      "B.Sc/B.Tech in CS, CUET-PG"),
            ("MBA",                                        "Master's",   "Business Administration","2 Years",   "₹1,35,000/year",   "Bachelor's Degree, CAT/MAT"),
            ("PhD in Hindi Literature",                    "PhD",        "Languages & Literature", "3-5 Years", "₹14,580/year",      "Master's Degree, NET/JRF"),
        ]
    },
    "U004": {
        "name":    "Jawaharlal Nehru University (JNU)",
        "country": "India",
        "city":    "New Delhi",
        "website": "https://www.jnu.ac.in",
        "courses": [
            ("M.A. in International Relations",            "Master's",   "Political Science",      "2 Years",   "₹3,168/year",       "Bachelor's Degree, CUET-PG"),
            ("M.A. in Economics",                          "Master's",   "Economics",              "2 Years",   "₹3,168/year",       "Bachelor's in Economics, CUET-PG"),
            ("M.Sc in Life Sciences",                      "Master's",   "Life Sciences",          "2 Years",   "₹2,748/year",       "B.Sc in Science, CUET-PG"),
            ("M.Sc in Biotechnology",                      "Master's",   "Biotechnology",          "2 Years",   "₹2,748/year",       "B.Sc in Life Sciences, GAT-B"),
            ("MBA",                                        "Master's",   "Business Administration","2 Years",   "₹1,20,000/year",   "Bachelor's Degree, CAT"),
            ("M.Phil in Environmental Studies",            "MPhil",      "Environmental Science",  "1 Year",    "₹2,748/year",       "Master's Degree"),
            ("PhD in Environmental Studies",               "PhD",        "Environmental Sciences", "3-5 Years", "₹2,748/year",       "Master's Degree, NET/JRF"),
        ]
    },
    "U005": {
        "name":    "Indian Institute of Technology Delhi (IIT Delhi)",
        "country": "India",
        "city":    "New Delhi",
        "website": "https://home.iitd.ac.in",
        "courses": [
            ("B.Tech in Computer Science & Engineering",   "Bachelor's", "Computer Science",       "4 Years",   "₹2,18,500/year",    "JEE Advanced, Class 12 with 75%+"),
            ("B.Tech in Civil Engineering",                "Bachelor's", "Civil Engineering",      "4 Years",   "₹2,18,500/year",    "JEE Advanced, Class 12 with 75%+"),
            ("B.Tech in Chemical Engineering",             "Bachelor's", "Chemical Engineering",   "4 Years",   "₹2,18,500/year",    "JEE Advanced, Class 12 with 75%+"),
            ("M.Tech in VLSI Design & Technology",         "Master's",   "Electronics Engineering","2 Years",   "₹60,000/semester",  "GATE qualified in EC/EE"),
            ("M.Tech in Computer Technology",              "Master's",   "Computer Science",       "2 Years",   "₹60,000/semester",  "GATE qualified in CS"),
            ("MBA",                                        "Master's",   "Business Administration","2 Years",   "₹3,50,000/year",   "Bachelor's Degree, CAT"),
            ("PhD in Biotechnology",                       "PhD",        "Biotechnology",          "4-6 Years", "₹16,800/semester",  "M.Tech/M.Sc + GATE/NET"),
        ]
    },
}


# ── BUILD DATASETS ────────────────────────────────────────────────────────────
def build_datasets():
    universities   = []
    courses        = []
    course_counter = 1

    for uid, udata in FALLBACK_DATA.items():

        universities.append({
            "university_id":   uid,
            "university_name": udata["name"],
            "country":         udata["country"],
            "city":            udata["city"],
            "website":         udata["website"],
        })

        live = scrape_shiksha(uid)
        time.sleep(2)  


        if len(live) >= 5:
            course_list = live
            print(f"  → LIVE data used for {uid}\n")
        else:
            course_list = udata["courses"]
            print(f"  → FALLBACK data used for {uid}\n")

        for course in course_list:
            if isinstance(course, dict):
                row = {
                    "course_id":     f"C{course_counter:03d}",
                    "university_id": uid,
                    "course_name":   course.get("course_name",  "Not Available"),
                    "level":         course.get("level",        "Not Available"),
                    "discipline":    course.get("discipline",   "Not Available"),
                    "duration":      course.get("duration",     "Not Available"),
                    "fees":          course.get("fees",         "Not Available"),
                    "eligibility":   course.get("eligibility",  "Not Available"),
                }
            else:
                row = {
                    "course_id":     f"C{course_counter:03d}",
                    "university_id": uid,
                    "course_name":   course[0],
                    "level":         course[1],
                    "discipline":    course[2],
                    "duration":      course[3],
                    "fees":          course[4],
                    "eligibility":   course[5],
                }
            courses.append(row)
            course_counter += 1

    df_uni     = pd.DataFrame(universities)
    df_courses = pd.DataFrame(courses)

    for col in ["fees", "duration", "eligibility", "discipline"]:
        df_courses[col] = df_courses[col].fillna("Not Available").replace("", "Not Available")

    df_uni     = df_uni.drop_duplicates(subset="university_id")
    df_courses = df_courses.drop_duplicates(subset="course_id")

    orphans = df_courses[~df_courses["university_id"].isin(df_uni["university_id"])]
    if len(orphans):
        print(f"WARNING: {len(orphans)} orphan courses found!")
    else:
        print(" Relational integrity OK")

    return df_uni, df_courses


def style_header(ws, num_cols, color):
    for col in range(1, num_cols + 1):
        cell            = ws.cell(row=1, column=col)
        cell.font       = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill       = PatternFill("solid", fgColor=color)
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border     = Border(
            bottom=Side(style="medium", color="FFFFFF"),
            right= Side(style="thin",   color="FFFFFF"),
        )
    ws.row_dimensions[1].height = 30


def style_rows(ws, num_rows, num_cols):
    for row in range(2, num_rows + 2):
        bg = "EAF2FB" if row % 2 == 0 else "FFFFFF"
        for col in range(1, num_cols + 1):
            cell           = ws.cell(row=row, column=col)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border    = Border(
                bottom=Side(style="thin", color="D5D8DC"),
                right= Side(style="thin", color="D5D8DC"),
            )
        ws.row_dimensions[row].height = 20


def set_col_widths(ws, df):
    for i, col in enumerate(df.columns, start=1):
        max_len = max(len(str(col)), df.iloc[:, i-1].astype(str).map(len).max())
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 50)


def export_excel(df_uni, df_courses, path):
    wb = Workbook()

    ws1 = wb.active
    if ws1 is None:
        ws1 = wb.create_sheet("Universities")
    ws1.title = "Universities"
    ws1.append(list(df_uni.columns))
    for row in df_uni.itertuples(index=False):
        ws1.append(list(row))
    style_header(ws1, len(df_uni.columns), "1A5276")
    style_rows(ws1, len(df_uni), len(df_uni.columns))
    set_col_widths(ws1, df_uni)
    ws1.freeze_panes = "A2"

    for r in range(2, len(df_uni) + 2):
        cell = ws1.cell(row=r, column=5)
        val = str(cell.value) if cell.value else ""
        if val.startswith("http"):
            cell.hyperlink = Hyperlink(ref=cell.coordinate, target=val)
            cell.font = Font(name="Arial", size=10, color="1F618D", underline="single")

    ws2 = wb.create_sheet("Courses")
    ws2.append(list(df_courses.columns))
    for row in df_courses.itertuples(index=False):
        ws2.append(list(row))
    style_header(ws2, len(df_courses.columns), "1E8449")
    style_rows(ws2, len(df_courses), len(df_courses.columns))
    set_col_widths(ws2, df_courses)
    ws2.freeze_panes = "A2"

    ws3 : Worksheet = wb.create_sheet("Summary")
    ws3["A1"] = "Dataset Summary"
    ws3["A1"].font = Font(bold=True, size=14, name="Arial", color="1A5276")
    ws3.row_dimensions[1].height = 30
    rows = [
        ("", ""),
        ("Total Universities",         "=COUNTA(Universities!A2:A100)"),
        ("Total Courses",              "=COUNTA(Courses!A2:A200)"),
        ("Avg Courses / University",   "=ROUND(COUNTA(Courses!A2:A200)/COUNTA(Universities!A2:A100),1)"),
        ("", ""),
        ("Countries",    ", ".join(df_uni["country"].unique())),
        ("Cities",       ", ".join(df_uni["city"].unique())),
        ("Course Levels","Bachelor's, Master's, MPhil, PhD"),
    ]
    for i, (label, value) in enumerate(rows, start=2):
        ws3.cell(row=i, column=1, value=label).font = Font(bold=True, name="Arial", size=11)
        ws3.cell(row=i, column=2, value=value).font = Font(name="Arial", size=11)
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 70

    wb.save(path)
    print(f"Saved: {path}")


def validate(df_uni, df_courses):
    print("\n" + "="*50)
    print("  VALIDATION REPORT")
    print("="*50)
    print(f"  Universities      : {len(df_uni)}")
    print(f"  Total Courses     : {len(df_courses)}")
    print(f"  Dup university_id : {df_uni['university_id'].duplicated().sum()}")
    print(f"  Dup course_id     : {df_courses['course_id'].duplicated().sum()}")
    orphans = df_courses[~df_courses["university_id"].isin(df_uni["university_id"])]
    print(f"  Orphan courses    : {len(orphans)}")
    print("-"*50)
    for uid in df_uni["university_id"]:
        count = len(df_courses[df_courses["university_id"] == uid])
        name  = df_uni[df_uni["university_id"] == uid]["university_name"].values[0]
        print(f"  {uid} | {count} courses | {name}")
    print("="*50 + "\n")


if __name__ == "__main__":
    print("="*50)
    print("  Indian University & Course Data Scraper")
    print("="*50 + "\n")

    os.makedirs("output", exist_ok=True)

    df_uni, df_courses = build_datasets()
    validate(df_uni, df_courses)
    export_excel(df_uni, df_courses, path="/mnt/user-data/outputs/university_data.xlsx")

    print("Done! Open output/university_data.xlsx")